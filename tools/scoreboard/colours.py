# -*- coding: utf-8 -*-
"""
Created on Mon Aug 29 14:55:20 2016

@author: badeape
"""
import xlsxwriter
from config.parameters import *
from config.config_scoreboard import *
from util.download_data import * # David's tool
import re
import openpyxl


# Data: Datafreme: data,  Scores in columns scoreL and scoreD
def  getColorGroup(row):
    if row['scoreL'] > 1 and row['scoreD']>-1:
        return 'red'
    elif ( (0.5 <row['scoreL'] <= 1)  and row['scoreD'] > -1) or ((-0.5< row['scoreL'] <= 0.5) and row['scoreD']>1):
        return 'orange'
    elif (row['scoreL'] > 0.5 and row['scoreD'] <= -1) :
        return 'yellow'
    elif (row['scoreL'] <= -0.5 and row['scoreD'] > 1) :
        return 'blue'
    elif ( -1 <row['scoreL'] <=-0.5  and row['scoreD'] <=1) or ( -0.5< row['scoreL'] <= 0.5 and row['scoreD']<=-1):
        return 'green'
    elif row['scoreL'] <= -1 and row['scoreD']<=1:
        return 'dark_green'
    elif row['scoreL'] <= 0.5 and row['scoreL'] > -0.5 and row['scoreD']<=1 and row['scoreD'] > -1:
        return 'white'

def getException(row):
    if (-0.5 < row['scoreL'] <= 0.5) and row['scoreD'] >=1 and row['ydiff']>=0 and row['sense']=='pos':
        return 'white'
    elif (-0.5 < row['scoreL'] <= 0.5) and row['scoreD'] >=1 and row['ydiff']<=0 and row['sense']=='neg':
        return 'white'
        
    elif (-1 < row['scoreL'] <= -0.5) and row['scoreD'] >=1 and row['ydiff']>=0 and row['sense']=='pos':
        return 'green'
    elif (-1 < row['scoreL'] <= -0.5) and row['scoreD'] >=1 and row['ydiff']<=0 and row['sense']=='neg':
        return 'green'
        
    elif row['scoreL'] <= -1 and row['scoreD'] >=1 and row['ydiff']>=0 and row['sense']=='pos':
        return 'dark_green'
    elif row['scoreL'] <= -1 and row['scoreD'] >=1 and row['ydiff']<=0 and row['sense']=='neg':
        return 'dark_green'
    else:
        return row['colour']
    
 # Create a workbook and add a worksheet.
workbook = xlsxwriter.Workbook(localpath+"COLOURS.xlsx")
worksheet = workbook.add_worksheet('All')

# Styles
# Format for the background colours 
red = workbook.add_format({'bg_color': '#FF0000', 'border':1, 'text_wrap': True})
orange = workbook.add_format({'bg_color': '#FF9900', 'border':1, 'text_wrap': True})
yellow = workbook.add_format({'bg_color': '#FFFF00', 'border':1, 'text_wrap': True})
blue = workbook.add_format({'bg_color': '#00CCFF', 'border':1, 'text_wrap': True})
white = workbook.add_format({'border':1, 'text_wrap': True})
green = workbook.add_format({'bg_color': '#66FF66', 'border':1, 'text_wrap': True})
dark_green = workbook.add_format({'bg_color': '#11BB11', 'border':1, 'text_wrap': True})

#Widths and heights
worksheet.set_column(1, 1, 25)
              
data=pd.read_csv(localpath+"SCORES.csv")     
data['colour']=data.apply(getColorGroup, axis=1) 
data['colour1']=data.apply(getException, axis=1)
data['ind'] = 'ID'+data.Order.astype(str)
data = data[['ind','indicator','geo','year','level','scoreL','ydiff','scoreD','category','LabelL','LabelD','colour','colour1']]
data.to_csv(localpath+'SCATTER_check.csv',index=False, float_format='%.3f')   
data = data[['ind','indicator','geo','year','level','scoreL','ydiff','scoreD','category','LabelL','LabelD','colour1']]
data.columns=['ind','indicator','geo','year','level','scoreL','ydiff','scoreD','category','LabelL','LabelD','colour']
data.to_csv(localpath+'SCATTER.csv',index=False, float_format='%.3f')

###### ------ Additional file to feed into worksheet "Input data" in F1's SCF Tables Excel file ------
print('Preparing data for file5...')
data_for_SCF=pd.read_csv(localpath+"SCORES_all_years.csv")
data_for_SCF['colour']=data_for_SCF.apply(getColorGroup, axis=1)
data_for_SCF['colour1']=data_for_SCF.apply(getException, axis=1)
data_for_SCF['ind'] = 'ID'+data_for_SCF.Order.astype(str)
max_year_per_ind = data_for_SCF.groupby('ind')['year'].transform(max)
data_for_SCF = data_for_SCF[~((data_for_SCF['ind'] == 'ID0') & # drop the fake year -- see comment "Artificially make the second-latest-available year..." in scoreboard_y.py
                              (data_for_SCF['year'] != data_for_SCF[data_for_SCF['ind']=='ID0']['year'].max()))]
data_for_SCF = data_for_SCF[['ind','indicator','geo','year','level','scoreL','ydiff','scoreD','category','LabelL','LabelD','colour1']]
data_for_SCF.columns=['ind','indicator','geo','year','level','scoreL','ydiff','scoreD','category','LabelL','LabelD','colour']
data_for_SCF = data_for_SCF[['ind','indicator','geo','year','colour']]
colour_to_int = {
    'red': 1, # Critical situations
    'orange': 2, # To watch
    'yellow': 3, # Weak but improving
    'white': 5, # On average
    'blue': 4, # Good but to monitor
    'green': 6, # Better than average
    'dark_green': 7 # Best performers
}
data_for_SCF['colour_num'] = data_for_SCF['colour'].map(colour_to_int)
data_for_SCF['ID_int'] = data_for_SCF['ind'].str.extract('(\d+)').astype(int)
reshaped_data_for_SCF = data_for_SCF.pivot_table(
    index=['ID_int', 'indicator', 'year'],
    columns='geo',
    values='colour_num',
    aggfunc='sum'
).reset_index()
def split_description(description):
    match_parentheses = re.search(r'(.*)\s(\(.*\))', description) # Check if the description contains suffix/details in parentheses
    match_index = re.search(r'(.*)\sIndex', description) # Check if the description contains the word "Index"
    match_age_cohort = re.search(r'(.*)\s15-29 age cohort', description) # Check if the description contains the phrase "15-29 age cohort"
    if match_parentheses:
        return match_parentheses.groups()
    elif match_index:
        return match_index.group(1), 'Index'
    elif match_age_cohort:
        return match_age_cohort.group(1), '15-29 age cohort'
    else:
        return description, ""
reshaped_data_for_SCF['Indicator'], reshaped_data_for_SCF['Details'] = \
    zip(*reshaped_data_for_SCF['indicator'].apply(split_description))
reshaped_data_for_SCF.sort_values(by=['year','ID_int'], inplace=True)
country_order_in_columns = ['BE','BG','CZ','DK','DE','EE','IE','EL','ES','FR','HR','IT','CY','LV','LT','LU','HU',
                            'MT','NL','AT','PL','PT','RO','SI','SK','FI','SE']
new_column_order = ['Indicator','Details','year'] + country_order_in_columns
reshaped_data_for_SCF[new_column_order].to_csv(localpath+'For_SCF_tables.csv',
                                               index=False)

df = reshaped_data_for_SCF[new_column_order]
# The ordered list of indicators
# TO BE ADJUSTED FOR EACH ANNUAL JER EDITION:
latest_years = {
    "Adult participation in learning" : 2022,
    "Early leavers from education and training" : 2022,
    "Share of individuals who have basic or above basic overall digital skills" : 2023,
    "Young people not in employment, education or training" : 2022,
    "Gender employment gap" : 2023,
    "Income quintile ratio" : 2022,
    "Employment rate" : 2023,
    "Unemployment rate" : 2023,
    "Long-term unemployment rate" : 2023,
    "Gross disposable household income (GDHI) per capita growth" : 2022,
    "At risk of poverty or social exclusion (AROPE) rate" : 2022,
    "At risk of poverty or social exclusion (AROPE) rate for children" : 2022,
    "Impact of social transfers (other than pensions) on poverty reduction" : 2022,
    "Disability employment gap" : 2022,
    "Housing cost overburden" : 2022,
    "Children aged less than 3 years in formal childcare" : 2022,
    "Self-reported unmet need for medical care" : 2022
}
# Initialize a dictionary to hold the dataframes for each country
country_dfs = {}
countries = df.columns[3:]  # List of country codes
# For each country, create a dataframe with the specified indicators and the last three available years
for country in countries:
    # Pivot the dataframe for the specific country with indicators as index and years as columns
    df_subset = df[['Indicator', 'year', country]].copy()
    df_subset['max_year_per_indicator'] = df_subset['Indicator'].map(latest_years)
    df_subset['year'] = df_subset['year'] - df_subset['max_year_per_indicator']
    df_subset = df_subset[df_subset['year'].isin([0, -1, -2])]
    df_subset['max_year_per_indicator'] = df_subset['Indicator']
    df_subset_pivot = df_subset.pivot(index='Indicator', columns='year', values=country)
    # Reindex the dataframe to ensure all required indicators are included in the appropriate order, filling missing values with 0
    df_country = df_subset_pivot.reindex(list(latest_years.keys())).fillna(0)
    df_country['0 = ...'] = df_country.index.map(latest_years)
    # Add the resulting dataframe to the dictionary
    country_dfs[country] = df_country
# Create an Excel writer object using pandas
excel_writer = pd.ExcelWriter(localpath+'For_SCF_tables_Input_Data_worksheet.xlsx')
# Write each dataframe to a separate sheet in the Excel file
for country, country_df in country_dfs.items():
    # Write the country dataframe to a sheet named after the country code
    country_df.to_excel(excel_writer, sheet_name=country)
# Save the Excel file
excel_writer.close()
## Doesn't work correctly
# source_wb = openpyxl.load_workbook(localpath+'For_SCF_tables_Input_Data_worksheet.xlsx')
# target_wb = openpyxl.load_workbook('U:\\04 Data and tools\\Reports\\scoreboard\\JER 25 SCF Tables 27032024.xlsx')
# for sheet_name in source_wb.sheetnames:
#     # Get the sheet from source workbook
#     source_sheet = source_wb[sheet_name]
#     # Create a new sheet in the target workbook with the same name
#     if sheet_name not in target_wb.sheetnames:
#         target_sheet = target_wb.create_sheet(sheet_name)
#     else:
#         target_sheet = target_wb[sheet_name]
#     # Copying the cell values from source to target
#     for row in source_sheet:
#         for cell in row:
#             target_sheet[cell.coordinate].value = cell.value
# # Save the target workbook
# target_wb.save(basicpath+'dissemination\\'+'Social Scoreboard - version 24 March - file5.xlsx')
print('Finished data prep for file5')
###### -----------------------------------------------------------------------------------------------
    
worksheet.write(1, 1,     "Best performers", dark_green) 
worksheet.write(2, 1,     "Better than average", green)
worksheet.write(3, 1,     "Good but deteriorating", blue)
worksheet.write(4, 1,     "On average", white )
worksheet.write(5, 1,     "Weak but improving", yellow) 
worksheet.write(6, 1,     "To watch", orange) 
worksheet.write(7, 1,     "Critical situations", red) 

  
myCol=2
for myIndicator in set(data['ind']): 
    myRow=1
    worksheet.write(0, myCol, myIndicator, white) 
    for myColor in ['dark_green', 'green', 'blue', 'white', 'yellow', 'orange', 'red']:
        countries=str(list(data[(data.colour==myColor) & (data.ind==myIndicator)]['geo']))
        countries=countries.replace("[","")
        countries=countries.replace("]","")
        countries=countries.replace("'","")
        worksheet.write(myRow, myCol, countries , eval(myColor))
        myRow+=1
    myCol+=1


for myIndicator in set(data['ind']): 
    worksheet = workbook.add_worksheet(myIndicator)
    worksheet.write(1, 1,     myIndicator) 
    worksheet.write(2, 1,     "Very high") 
    worksheet.write(3, 1,     "High") 
    worksheet.write(4, 1,     "On average") 
    worksheet.write(5, 1,     "Low")
    worksheet.write(6, 1,     "Very low") 
    worksheet.write(1, 2,     "Much higher than average") 
    worksheet.write(1, 3,     "Higher than average") 
    worksheet.write(1, 4,     "On average") 
    worksheet.write(1, 5,     "Lower than average") 
    worksheet.write(1, 6,     "Much lower than average")
        
    myRow=2
    for Row in ['Very low','Low','On average','High','Very high']:
        myCol=2
        for Col in ['Much lower than average','Lower than average','On average','Higher than average','Much higher than average']:
            countries=str(list(data[(data.LabelL==Row) & (data.LabelD==Col) & (data.ind==myIndicator)]['geo']))
            countries=countries.replace("[","")
            countries=countries.replace("]","")
            countries=countries.replace("'","")
            worksheet.write(myRow, myCol, countries)
            myCol+=1
        myRow+=1

       
workbook.close()

##############################################################################################
# Added in August 2023 by Alek -- replaces old parts of macro "format_diss02" from "Scoreboard macros.xlsm"

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side

# Helper
def copy_style(src_cell, dest_cell):
    # Copy font
    font = src_cell.font
    new_font = Font(name=font.name, sz=font.sz, color=font.color,
                    bold=font.bold, italic=font.italic, vertAlign=font.vertAlign,
                    underline=font.underline, strike=font.strike, scheme=font.scheme)
    dest_cell.font = new_font
    # Copy border
    border = src_cell.border
    new_border = Border(left=border.left, right=border.right, top=border.top,
                        bottom=border.bottom, diagonal=border.diagonal,
                        diagonal_direction=border.diagonal_direction,
                        outline=border.outline, start=border.start, end=border.end)
    dest_cell.border = new_border
    # Copy alignment
    alignment = src_cell.alignment
    new_alignment = Alignment(horizontal=alignment.horizontal, vertical=alignment.vertical,
                              text_rotation=alignment.text_rotation, wrap_text=alignment.wrap_text,
                              shrink_to_fit=alignment.shrink_to_fit, indent=alignment.indent)
    dest_cell.alignment = new_alignment
    # Copy fill
    fill = src_cell.fill
    new_fill = PatternFill(fill_type=fill.fill_type, start_color=fill.start_color,
                           end_color=fill.end_color, fgColor=fill.fgColor)
    dest_cell.fill = new_fill
    # Copy number format (this is directly assignable)
    dest_cell.number_format = src_cell.number_format
    # Copy protection
    protection = src_cell.protection
    new_protection = Protection(locked=protection.locked, hidden=protection.hidden)
    dest_cell.protection = new_protection

# Import worksheet "All" and transpose
current_wb = openpyxl.load_workbook(localpath+"COLOURS.xlsx")
current_sheet = current_wb['All']
new_sheet = current_wb.create_sheet(title='All (transposed)',index=0) # 0 = at the first position
for row in range(1, current_sheet.max_row + 1):
  for col in range(1, current_sheet.max_column + 1):
    old_cell = current_sheet.cell(row, col)
    new_cell = new_sheet.cell(col, row)
    new_cell.value = old_cell.value
    copy_style(old_cell, new_cell)

# Sort by IDs
new_sheet.delete_rows(1) # Empty
new_sheet.cell(1, 1).value ='ID-100' # needed for sorting
all_rows = [row for row in new_sheet.iter_rows(values_only=True)]
#sorted_rows = sorted(all_rows, key=lambda row: int(row[0][2:]))
all_rows_sorted = sorted(all_rows, key=lambda row: int(row[0][2:]))
for row in new_sheet.iter_rows(): # clean-up
    for cell in row:
        cell.value = None
for row_index, row_items in enumerate(all_rows_sorted, 1):  # 1-based indexing for worksheets
    for col_index, cell_value in enumerate(row_items, 1):
        new_sheet.cell(row=row_index, column=col_index, value=cell_value)

# Replace ID codes with indicator names
correspondence_dict = pd.read_csv(localpath+"SCORES.csv",
                                  usecols=["code","indicator"]).drop_duplicates().set_index('code')['indicator'].to_dict()
correspondence_dict['ID-100'] = ""
for row in new_sheet.iter_rows(min_col=1, max_col=1, values_only=False):  # Ensure values_only = False
    cell = row[0]
    if cell.value in correspondence_dict:
        cell.value = correspondence_dict[cell.value]

# Insert column with the 3 broad indicator categories
new_sheet.insert_cols(idx=1)
beige_fill = PatternFill(start_color="F5F5DC", end_color="F5F5DC", fill_type="solid")
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
new_sheet.merge_cells('A2:A7')
new_sheet['A2'].value = "Equal opportunities"
new_sheet['A2'].fill = beige_fill
new_sheet.merge_cells('A8:A11')
new_sheet['A8'].value = "Fair working conditions"
new_sheet['A8'].fill = beige_fill
new_sheet.merge_cells('A12:A18')
new_sheet['A12'].value = "Social protection and inclusion"
new_sheet['A12'].fill = beige_fill

# Add the column with years
indics_years = pd.read_csv(localpath+"SCORES.csv",
                           usecols=["code", "year"]).drop_duplicates()
indics_years['code'] = pd.Categorical(indics_years['code'],
                                      categories=list(correspondence_dict.keys()), ordered=True)
indics_years['numeric_code'] = indics_years['code'].str.lstrip('ID').astype(int)
indics_years = indics_years.sort_values(by='numeric_code').drop(columns='numeric_code')
new_sheet.insert_cols(idx=3)
for index, value in enumerate(indics_years['year'], start=2):
    new_sheet.cell(row=index, column=3).value = value

# Apply styles
bold_font = Font(bold=True)
centered_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
for col in range(1, 4):  # Columns are 1-indexed in openpyxl
    for row in range(1, new_sheet.max_row + 1):
        cell = new_sheet.cell(row=row, column=col)
        cell.font = bold_font
        cell.alignment = centered_alignment
        cell.border = thin_border
new_sheet.column_dimensions['B'].width = 35 # Indicator names
new_sheet.column_dimensions['C'].width = 7 # Years
for l in ['A', 'D','E','F','G','H','I','J']:
    new_sheet.column_dimensions[l].width = 20

new_sheet.sheet_view.zoomScale = 80

current_wb.remove(current_sheet)
new_sheet.title = "All"
current_wb.save(localpath+"COLOURS.xlsx")
